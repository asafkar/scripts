#!/tools/bin/python3.6
import Classes
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import xlrd
import math

# import multiprocessing
import settings


def help():
	print("----------Specman Register files generator------------")
	print("Usage: ")
	print("specman_register_gen.py <top_excel> <tag>")
	print("<top_excel> is the excel of the registers, mandatory")
	print("if <tag> isn't specific, latest will be used")
	print("------------------------------------------------------")

""" This function will convert a filename to the relevant 
filename in the REVISIONS directory - thus using the 
defined TAG """


def file_name_from_tag(full_filename):
	if settings.tag_used is None or full_filename == settings.local_excel_filename:  # don't check the given excel for tag
		return full_filename
	else:
		block_name = full_filename.split("/")[-3]  # get the name of the block
		file_name= full_filename.split("/")[-1]
		dir_before_file=full_filename.split("/")[-2]

		if block_name in settings.tag_dict:
			return settings.revisions_dir+settings.tag_dict[block_name] + "/" + dir_before_file + "/" + file_name
		else:
			return full_filename


""" Read 'spaces' tab in the excel book,
get all spaces to the Spaces_list """

def file_to_wb(filename):
	fn=file_name_from_tag(filename)
	if fn.endswith("xls"):
		cvt_xls_to_xlsx(fn, "temp.xlsx")
		fn = "temp.xlsx"
	return load_workbook(filename=fn, data_only=True)


# a wb might have a spaces tab, and it might not
def parse_wb(wb, wb_name="", addr_offset="0", name_prefix=""):
	spaces_list = []
	chip_registers_list = []

	reached_data = False

	if 'Spaces' in wb.sheetnames and not wb_name:
		for row in tuple(wb["Spaces"].rows):

			# skip top part of excel sheet, usually not relevant
			if reached_data is False:
				for cell in row:
					if cell.value == "type":
						reached_data = True
						break  # goto next row/line (with actual data), and look at whole row
			else:
				# print(cell.value, end=" ") if cell.value else "" # print only non empty cells
				if (row[2].value is None) or (row[0].value != "reg_block" and row[0].value != "reg_space" and row[0].value != "ext_block"): ##ak fixme extblock???
					continue
				spaces_list.append(Classes.Spaces(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value))  # type,name,space_addr,bits,description):
				# pprint(vars(Spaces_list[-1]))
				# print(" ... ")
	else:
		if wb_name in wb.sheetnames:
			# print("adding ", wb_name, " to spaces list")
			spaces_list.append(Classes.Spaces("reg_block", wb_name, addr_offset, "32", ""))  # type,name,space_addr,bits,description):
		else:
			raise NameError

	""" Parse all spaces in Spaces_list - 
	go over each one, open the specific tab, and add the tab's content 
	to the register_list"""
	for space in spaces_list:
		if space.space_type == "remark":
			continue
		# print("Iterating space " + space.name)
		if space.space_addr.startswith("0x"):
			space.space_addr = space.space_addr.split("x")[1]
		size_of_addr = len(space.space_addr)  # num of bytes
		# chip_registers_list = [*chip_registers_list, *parse_tab(wb, space.name, int(bin(int(str(space.space_addr), 16))[size_of_addr*4-int(space.bits)+2:size_of_addr*4+2], 2) << (32 - int(space.bits)), name_prefix)]
		chip_registers_list = [*chip_registers_list, *parse_tab(wb, space.name, int(str(space.space_addr), 16) << (32 - int(size_of_addr)*4), name_prefix)]
	return chip_registers_list


# remove white spaces
def s(in_str):
	if in_str is not None:
		return in_str.strip()
	else:
		return ""


""" This function receives a list of rows (in a tab, or external tab),
which it parses, and returns a list of registers (w/o offset)"""


def parse_tab(curr_wb, space_name, tab_addr_offset, name_prefix=""):
	if name_prefix == "None" or name_prefix == "":
		name_prefix_fixed = ""
	elif name_prefix.lower().endswith("none"):  # peel off 'none' endings if some exist
		name_prefix_fixed = name_prefix[0:-4]
	elif name_prefix is not None:
		name_prefix_fixed = name_prefix + "_"

	rows_list=curr_wb[space_name].rows
	tab_registers = []  # current register list
	# rows_iterator = iter(tuple(rows_list))  # iterator that goes through the list of rows in current tab being processed
	rows_iterator = iter(rows_list)  # iterator that goes through the list of rows in current tab being processed
	# for row in rows_iterator:  # go over all rows in specific space tab
	done_iter = False
	temp_dont_advance = False
	while not done_iter:
		try:
			if not temp_dont_advance:
				row = next(rows_iterator)
			else:
				temp_dont_advance = False
		except StopIteration:  # if reached end of sheet
			done_iter = True
		else:
			# print("DEBUG Parsing " +str(row[0].value) + " " + str(row[2].value) + " " + str(row[3].value))
			try:
				curr_reg_name = str(row[3].value)
			except:
				curr_reg_name = ""

			if (row[0].value is None) or (row[0].value == "Remark") or (row[0].value == "None"):
				continue
			if s(row[0].value) == "module":
				# print("parsing module ", row[1].value)
				registers_list_temp = parse_wb(curr_wb, row[1].value, str(hex(int(str(row[2].value), 16)+tab_addr_offset)), name_prefix_fixed+curr_reg_name)
				tab_registers = [*tab_registers, *registers_list_temp]  				# append list to existing list
				# print("finished parsing module ", row[1].value)

			if s(row[0].value) == "fmodule":
				# print("opening ", row[4].value, " for fmodule ", row[1].value)
				wb_temp = file_to_wb(row[4].value)
				registers_list_temp = parse_wb(wb_temp, row[1].value, str(hex(int(str(row[2].value), 16)+tab_addr_offset)), name_prefix_fixed + curr_reg_name)
				tab_registers = [*tab_registers, *registers_list_temp]  				# append list to existing list

			if s(row[0].value) == "wide_reg" or s(row[0].value) == "wide_ro":
				N = int(row[5].value) # num of fields
				M = int(row[3].value) # num of bits
				fields_in_reg = 32/M
				bits_used_in_each_reg = M*fields_in_reg
				total_num_of_regs = math.ceil(N*M/bits_used_in_each_reg)

				rw_attribute = "RW0"
				if s(row[0].value) == "wide_ro":
					rw_attribute = "RO0"


				for ii in range(0, total_num_of_regs):
					temp_reg = Classes.Register(settings.regs_prefix + "_" + name_prefix_fixed + row[1].value + "_"+str(ii),
												str(int(str(row[2].value), 16) + tab_addr_offset+ii*4),
												bits_used_in_each_reg, rw_attribute)
					temp_field = Classes.Field(row[1].value, "[31:0]", 0, "") ## wide reg doesn't have fields, create empty ones
					temp_reg.add_fields(temp_field)
					tab_registers.append(temp_reg)

			if s(row[0].value) == "register":
				temp_reg = Classes.Register(settings.regs_prefix + "_" + name_prefix_fixed+row[1].value, str(int(str(row[2].value), 16) + tab_addr_offset),
											row[3].value, row[4].value)
				# if "none" in temp_reg.name:
				# 	print (temp_reg.name)
				# 	print(name_prefix_fixed)
				parse_fields = True
				if s(row[4].value) == "RP1": ## RP1 doesn't have fields
					parse_fields = False
					temp_field = Classes.Field(row[1].value, "0", 0, "")
					temp_reg.add_fields(temp_field)
					parse_fields = False
				while parse_fields:
					try:
						row = next(rows_iterator)  # goto next row
						if s(row[0].value) != "field":
							parse_fields = False

							# if the parse_field method parsed too much and there's no space in the excel, don't go further
							curr_row0_val = s(row[0].value)
							if (curr_row0_val == "fmodule" or curr_row0_val == "module" or curr_row0_val == "eblock" \
									or curr_row0_val == "fblock" or curr_row0_val == "register" or curr_row0_val == "wide_reg"):
								temp_dont_advance = True

						else:
							temp_field = Classes.Field(row[1].value, row[3].value, row[4].value, row[6].value)
							temp_reg.add_fields(temp_field)
					except StopIteration:  # if reached end of sheet
						done_iter = True
						parse_fields = False
				temp_reg.sort_fields()
				tab_registers.append(temp_reg)
	# for reg in tab_registers:
		# pprint(reg.name + "  addr: " + hex(int(str(reg.addr))))
	return tab_registers


def cvt_xls_to_xlsx(src_file_path, dst_file_path):
	book_xls = xlrd.open_workbook(src_file_path)
	book_xlsx = Workbook()

	sheet_names = book_xls.sheet_names()
	for sheet_index in range(0, len(sheet_names)):
		sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
		if sheet_index == 0:
			sheet_xlsx = book_xlsx.active
			sheet_xlsx.title = sheet_names[sheet_index]
		else:
			sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

		for row in range(0, sheet_xls.nrows):
			for col in range(0, sheet_xls.ncols):
				sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

	book_xlsx.save(dst_file_path)
